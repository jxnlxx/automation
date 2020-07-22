# -*- coding: utf-8 -*-
"""
Created on Mon Jun  8 15:27:24 2020

@author: JLee35
"""


import os
import datetime
import requests
import openpyxl
import pandas as pd
import xlsxwriter as xl
from xlsxwriter.utility import xl_range

# generate timestamp for 'start_time' so section at end doesnt break - remove this when copying to healthchecks_full

start_time = datetime.datetime.now().replace(microsecond=0)

# generate timestamp for after deepcrawl section of script
mid_time = datetime.datetime.now().replace(microsecond=0)


# load healthcheck list

hc_list = pd.read_excel( r'L:\Commercial\Operations\Technical SEO\Technical Healthchecks\Healthcheck List.xlsx', sheet_name = 'Healthcheck List' )

#iterate through list to grab data, ping server to get data, then update reports


# =============================================================================
# create dates 
# =============================================================================

report_month = datetime.datetime.now()
#year = report_month.strftime( '%Y' )
#month = report_month.strftime( '%m %b')
report_month = report_month.strftime( '%b-%y' )


# iterate through rows of healthcheck list and produce one healthcheck per client

for i in hc_list.index:
    client = hc_list['Client'][i]
    project_id = hc_list['Project ID'][i]
    brand = hc_list[ 'Brand' ][i]
        
# load workbook as openpyxl and convert to dataframe 
# (gets around an issue where pandas wouldn't load the workbook properly)
    
    print( f'\nOpening {client} Healthcheck')
    try:
        df = openpyxl.load_workbook( filename=fr'L:\Commercial\Operations\Technical SEO\Automation\Technical Healthchecks\Data\Crawl Data\{client} - Tech Healthcheck Template.xlsx', data_only=True )
        df = df[ 'Healthcheck' ]
        df = pd.DataFrame( df.values )
        df_rows, df_cols = df.shape
        xl_rows = df_rows + 4
        print( f'Starting to Write {client} Healthcheck' )
    except FileNotFoundError:
        print( f'{client} Healthcheck Template not found' )
        continue    
    df_rows, df_cols = df.shape
    xl_rows = df_rows+4
    df = df.fillna( '' )

# create blank workbook with xlsxwriter

#    wb = xl.Workbook( fr'L:\Commercial\Operations\Technical SEO\Technical Healthchecks\01. Healthchecks\{year}\{month}\{client} - Tech Healthcheck.xlsx' )
    wb = xl.Workbook( fr'L:\Commercial\Operations\Technical SEO\Technical Healthchecks\01. Healthchecks\{client} - Tech Healthcheck.xlsx' )
    ws = wb.add_worksheet( 'Healthcheck' )
    
# column widths

    ws.set_column( 0, 0, 5 )
    ws.set_column( 1, 1, 45 )
    ws.set_column( 2, 2, 90 )
    ws.set_column( 3, 3, 60 )

# brand colours

    ipro_colour = '#8dc63f'
    ipro_spark = '#066bb1'

    carat_colour = '#00beff'
    carat_spark = '#4ee6C6'

    vizeum_colour = '#fac600'
    vizeum_spark = '#f6861f'

    text = '#636363'

# additional formatting

    ws.freeze_panes(6, 2)                                                      # Freeze first 6 rows and first 2 columns.
    ws.set_zoom(70)                                                            # Set zoom to 70%
    ws.hide_gridlines(2)                                                       # Hide Gridlines
#    ws.hide_row_col_headers()                                                  # hide headers - not great for unhiding hiddent tabs
    
#  CELL FORMATTING

# heading formatting (conditional on brand client is with)

    if brand == 'Carat':
        client_name = wb.add_format({ 'bold':True, 'indent':0, 'font_color':text, 'font_size':20, 'align':'center', 'valign':'vcenter', 'rotation':0 })
        l_head = wb.add_format({ 'bold':False, 'indent':0, 'bg_color':carat_colour, 'font_color':carat_colour, 'font_size':12, 'align':'center', 'valign':'vcenter', 'rotation':0, 'border_color':text, 'top':1, 'bottom':1, 'left':1, 'right':0 })
        c_head = wb.add_format({ 'bold':False, 'indent':0, 'bg_color':carat_colour, 'font_color':carat_colour, 'font_size':12, 'align':'center', 'valign':'vcenter', 'rotation':0, 'border_color':text, 'top':1, 'bottom':1, 'left':0, 'right':0 })
        r_head = wb.add_format({ 'bold':False, 'indent':0, 'bg_color':carat_colour, 'font_color':'#ffffff', 'font_size':12, 'align':'center', 'valign':'vcenter', 'rotation':90, 'border_color':text, 'top':1, 'bottom':1, 'left':1, 'right':1, 'num_format':'mmm-yy' })

    if brand == 'Vizeum':
        client_name = wb.add_format({ 'bold':True, 'indent':0, 'font_color':text, 'font_size':20, 'align':'center', 'valign':'vcenter', 'rotation':0 })
        l_head = wb.add_format({ 'bold':False, 'indent':0, 'bg_color':vizeum_colour, 'font_color':vizeum_colour, 'font_size':12, 'align':'center', 'valign':'vcenter', 'rotation':0, 'border_color':text, 'top':1, 'bottom':1, 'left':1, 'right':0 })
        c_head = wb.add_format({ 'bold':False, 'indent':0, 'bg_color':vizeum_colour, 'font_color':vizeum_colour, 'font_size':12, 'align':'center', 'valign':'vcenter', 'rotation':0, 'border_color':text, 'top':1, 'bottom':1, 'left':0, 'right':0 })
        r_head = wb.add_format({ 'bold':False, 'indent':0, 'bg_color':vizeum_colour, 'font_color':'#ffffff', 'font_size':12, 'align':'center', 'valign':'vcenter', 'rotation':90, 'border_color':text, 'top':1, 'bottom':1, 'left':1, 'right':1, 'num_format':'mmm-yy' })

    if brand not in [ 'Carat', 'Vizeum' ]:
        client_name = wb.add_format({ 'bold':True, 'indent':0, 'font_color':text, 'font_size':20, 'align':'center', 'valign':'vcenter', 'rotation':0 })
        l_head = wb.add_format({ 'bold':False, 'indent':0, 'bg_color':ipro_colour, 'font_color':ipro_colour, 'font_size':12, 'align':'center', 'valign':'vcenter', 'rotation':0, 'border_color':text, 'top':1, 'bottom':1, 'left':1, 'right':0 })
        c_head = wb.add_format({ 'bold':False, 'indent':0, 'bg_color':ipro_colour, 'font_color':ipro_colour, 'font_size':12, 'align':'center', 'valign':'vcenter', 'rotation':0, 'border_color':text, 'top':1, 'bottom':1, 'left':0, 'right':0 })
        r_head = wb.add_format({ 'bold':False, 'indent':0, 'bg_color':ipro_colour, 'font_color':'#ffffff', 'font_size':12, 'align':'center', 'valign':'vcenter', 'rotation':90, 'border_color':text, 'top':1, 'bottom':1, 'left':1, 'right':1, 'num_format':'mmm-yy' })

# table Formatting

    sub1 = wb.add_format({ 'bold':True, 'indent':1, 'font_color':text, 'font_size':18, 'align':'left', 'valign':'vcenter', 'rotation':0, 'border_color':text, 'top':1, 'bottom':1, 'left':0, 'right':0 })
    sub2 = wb.add_format({ 'bold':True, 'indent':1, 'font_color':'#ffffff', 'font_size':18, 'align':'left', 'valign':'vcenter', 'rotation':0, 'border_color':text, 'top':1, 'bottom':1, 'left':0, 'right':0 })
    issue = wb.add_format({ 'bold':True, 'indent':1, 'font_color':text, 'font_size':12, 'align':'left', 'valign':'vcenter', 'rotation':0, 'border_color':text, 'top':1, 'bottom':1, 'left':1, 'right':1 })
    detail = wb.add_format({ 'bold':False, 'indent':1, 'font_color':text, 'font_size':12, 'align':'left', 'valign':'vcenter', 'rotation':0, 'border_color':text, 'top':1, 'bottom':1, 'left':1, 'right':1 })
    spark = wb.add_format({ 'bold':False, 'indent':1, 'font_color':'#ffffff', 'font_size':12, 'align':'left', 'valign':'vcenter', 'rotation':0, 'border_color':text, 'top':1, 'bottom':1, 'left':1, 'right':1 })
    data = wb.add_format({ 'bold':False, 'indent':0, 'font_color':text, 'font_size':12, 'align':'center', 'valign':'vcenter', 'rotation':0, 'border_color':text, 'top':1, 'bottom':1, 'left':1, 'right':1 })

# row groups for adding formatting

    xl_head = [ 5 ]
    xl_sub = [ 6, 14, 21, 28, 37, 42, 46 ]
    xl_data = [ 7, 8, 9, 10, 11, 12, 13, 15, 16, 17, 18, 19, 20, 22, 23, 24, 25, 26, 27, 29, 30, 31, 32, 33, 34, 35, 36, 38, 39, 40, 41, 43, 44, 45, 47, 48, 49, 50, 51, 52, 53 ]
    
# format + fill table header
    
    ws.set_row( 2, 30 )
    ws.write(2, 2, f'{client} Healthcheck', client_name)
#    print('1')
# format + fill table data
    
    x = 5
    n = 0 #for df.loc[n]
    while x <= xl_rows:
#        print('2')
        if x in xl_head:
#            print('3')
            y = 1
            df_list = list(df.iloc[n])
            ws.set_row ( x, 60 )
            while y <= df_cols:
                if y == 1:
                    ws.write( x, y, df_list[y-1], l_head)
                    y += 1
                if y == 2:
                    ws.write( x, y, df_list[y-1], c_head)
                    y += 1
                if y == 3:
                    ws.write( x, y, '', c_head)
                    y += 1
                if y in range ( 4, df_cols ):
                    ws.write( x, y, df_list[y-1], r_head)
                    y += 1
                if y == df_cols:
                    ws.write( x, y, df_list[y-1], r_head)
                    y += 1
                    x += 1
                    n += 1
        if x in xl_sub: 
#            print('4')
            y = 1
            df_list = list(df.iloc[n])
            ws.set_row( x, 60 )
            while y <= df_cols:
                if y == 1:
                    ws.write( x, y, df_list[y-1], sub1)
                    y += 1
                if y in range(2, df_cols):
                    ws.write( x, y, df_list[y-1], sub2)
                    y += 1
                if y == df_cols:
                    ws.write( x, y, df_list[y-1], sub2)
                    y += 1
                    x += 1
                    n += 1
        if x in xl_data:
#            print('5')
            y = 1
            df_list = list(df.iloc[n])
            ws.set_row( x, 30 )
            while y <= df_cols:
                if y == 1:
                    ws.write( x, y, df_list[y-1], issue)
                    y += 1
                if y == 2:
                    ws.write( x, y, df_list[y-1], detail)
                    y += 1
                if y == 3:
                    ws.write( x, y, '', spark)
                    y += 1
                if y in range ( 4, df_cols ):
                    ws.write( x, y, df_list[y-1], data)
                    y += 1
                if y == df_cols:
                    ws.write( x, y, df_list[y-1], data)
                    y += 1
                    x += 1
                    n += 1        

# Add sparklines

    x, y = 7, 3

    if brand == 'Carat':
        while x <= xl_rows:
            sparkline = xl_range( x, 4, x, df_cols)
            ws.add_sparkline( x, y, { 'range': sparkline, 'series_color': carat_spark, 'markers':True } )
            x += 1
        ws.insert_image('B2', fr'L:\Commercial\Operations\Technical SEO\Automation\Technical Healthchecks\Data\Branding\Logos\Carat Logo.png', { 'y_offset' : 5, 'x_scale' : 0.13, 'y_scale': 0.13 } )
        ws.set_tab_color(carat_colour)

    if brand == 'Vizeum':
        while x <= xl_rows:
            sparkline = xl_range( x, 4, x, df_cols)
            ws.add_sparkline( x, y, { 'range': sparkline, 'series_color': vizeum_spark, 'markers':True } )
            x += 1
        ws.insert_image('B1', fr'L:\Commercial\Operations\Technical SEO\Automation\Technical Healthchecks\Data\Branding\Logos\Vizeum Logo.png', { 'x_offset' : 20, 'y_offset' : 10, 'x_scale' : 0.4, 'y_scale': 0.4 } )
        ws.set_tab_color(vizeum_colour)
    if brand not in [ 'Carat', 'Vizeum' ]:
        while x <= xl_rows:
            sparkline = xl_range( x, 4, x, df_cols)
            ws.add_sparkline( x, y, { 'range': sparkline, 'series_color': ipro_spark, 'markers':True } )
            x += 1
        ws.insert_image('B2', fr'L:\Commercial\Operations\Technical SEO\Automation\Technical Healthchecks\Data\Branding\Logos\iPro Logo.png', {'x_scale' : 0.22, 'y_scale': 0.22 } )
        ws.set_tab_color(ipro_colour)   

# hide columns

    if df_cols > 9:
        ws.set_column( 4, df_cols-6, 0 )
     
    wb.close()
    print( f'{client} Healthcheck Done'
           '\n'
           '\n*   *   *   *   *   *   *   *   *   *' )
#    break

print( 'DURN!')

end_time = datetime.datetime.now().replace(microsecond=0)

dc_time = mid_time - start_time
xl_time = end_time - mid_time
total_time = end_time - start_time


print(f'Time taken to get data from DeepCrawl: {dc_time}','\n'
      f'Time taken to generate Excel healthchecks: {xl_time}','\n'
      f'Total time taken: {total_time}')
