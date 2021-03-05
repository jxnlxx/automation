# -*- coding: utf-8 -*-
"""
Created on Fri Feb  7 12:16:23 2020

@author: Jon Lee
"""

import os
import datetime
import requests
import openpyxl
import pandas as pd
import xlsxwriter as xl
from xlsxwriter.utility import xl_range


# =============================================================================
# create dates 
# =============================================================================

report_month = datetime.datetime.now()
year = report_month.strftime( '%Y' )
month = report_month.strftime( '%m %b')
report_month = report_month.strftime( '%b-%y' )

# =============================================================================
# log in to deepcrawl
# =============================================================================

url = 'https://api.deepcrawl.com/sessions'
key = '1598'
value = 'FPXwkQUsX-qqqnhB_HLbHzO9brrycphhHKQwBwHOneNTP5Ur-5Dx_k2y_E57__4gWSGj810c'

# POST request to api.deepcrawl.com/sessions with the API key and value

response = requests.post( url, auth=( key, value ), verify=True )

# convert JSON string into dict

response_json = response.json()

# assign the 'token' in the dict to token variable

token = response_json[ 'token' ]

# create variable for adding into header of requests

headers = { 'X-Auth-Token' : token }

# print 'token'
print(token)






'''
# =============================================================================
# update healthcheck templates
# =============================================================================
'''






# load healthcheck list

hc_list = pd.read_excel( r'L:\Commercial\Operations\Technical SEO\Technical Healthchecks\Healthcheck List.xlsx', sheet_name = 'Healthcheck List' )

#iterate through list to grab data, ping server to get data, then update reports

for i in hc_list.index:
    n = 0
    client = hc_list['Client'][i]
    project_id = hc_list['Project ID'][i]
    if project_id == '':
        print( f'{client} does not have a project ID in the Healthchek List. Please rectify.')
        continue
    else:
        print( f'\nInitialising {client}' )
        pass
    
    brand = hc_list[ 'Brand' ][i]
    url = f'https://api.deepcrawl.com/accounts/117/projects/{project_id}'
    response = requests.get( url, headers=headers )

# handle errors if project does not exists on DeepCrawl

    if response.status_code != 200:
        print( '{client} Project does not exist on DeepCrawl - please check Project ID on Healthcheck List' )
        continue
    else:
        response_json = response.json()
    

# if no last crawl, skip client
    site_primary = response_json.get( 'site_primary' )
    last_crawl = response_json.get( '_crawls_last_href' )
    if last_crawl == None:
        print( f'{client} has no _crawls_last_href.')
        continue

# begin building dataframe for crawl

    print( f'Getting Data for {client}\nURL is {site_primary}\n')
    df = pd.DataFrame()                                                                 # create a blank DataFrame for the project
    url = 'https://api.deepcrawl.com' + last_crawl + '/reports?page=1&per_page=200'     # create a URL from the href 
    crawl = requests.get( url, headers=headers )                                        # ping the URL to get the data
    crawl_json = crawl.json()                                                           # convert the JSON response into something pythonic
    df = df.append( crawl_json, ignore_index = True )                                   # fill the DataFrame with data from the response
    n += 1                                                                              # add 1 to 'n'
    print( n, url )                                                                     # print 'n, url' so the script shows it's not idle
    while url != 'https://api.deepcrawl.com' + crawl.links[ 'last' ][ 'url' ]:          # check if the URL is the same as the last page in the response header
        url = 'https://api.deepcrawl.com' + crawl.links[ 'next' ][ 'url']               # if not, change the url to the 'next' page in the pagination
        crawl = requests.get( url, headers=headers )                                    # ping the new url to get the data
        crawl_json = crawl.json()                                                       # convert the JSON response into something pythonic
        df = df.append( crawl_json, ignore_index = True )                               # append the data to the DataFrame
        n += 1                                                                          # add 1 to 'n'
        print( n, url )                                                                 # print 'n, url' so the script shows it's not idle
            
# remove duplicates from 'report_template' column

    df = df.drop_duplicates( subset = 'report_template'  )                              # de-dupe the sreadsheet based on the 'report template' column

# filter columns for 'report_template' and 'basic_total'

    df = df[[ 'report_template', 'basic_total' ]]                                       # remove all columns from sheet except 'report_template' and 'basic_total' columns
#    df = df.rename( columns={ 'basic_total' : report_month })
# open template

# see if client's template exists, and if not, open blank template
    
    try:
        crawl_data = pd.read_excel( fr'L:\Commercial\Operations\Technical SEO\Automation\Technical Healthchecks\03. Data\Crawl Data\{client} - Tech Healthcheck Template.xlsx', sheet_name = 'Data' )
    except FileNotFoundError:
        crawl_data = pd.read_excel( fr'L:\Commercial\Operations\Technical SEO\Automation\Technical Healthchecks\03. Data\Blank Tech Healthcheck Template.xlsx', sheet_name = 'Data', read_only=True )

# merge the new data with the old data from the template
        
    crawl_data = crawl_data.merge( df, on= 'report_template', how='left' )
    crawl_data = crawl_data.rename( columns={ 'basic_total' : report_month })
    crawl_data = crawl_data.fillna( 0 )

    
# update 'data' tab


    
    df_rows, df_cols = crawl_data.shape


    x, y = 1, 1
    try:
        template = openpyxl.load_workbook( fr'L:\Commercial\Operations\Technical SEO\Automation\Technical Healthchecks\03. Data\Crawl Data\{client} - Tech Healthcheck Template.xlsx' )
    except FileNotFoundError:
        template = openpyxl.load_workbook( fr'L:\Commercial\Operations\Technical SEO\Automation\Technical Healthchecks\03. Data\Blank Tech Healthcheck Template.xlsx' )
        print( f'Creating New Template for {client}')
    data = template[ 'Data' ]
    healthcheck = template[ 'Healthcheck' ]

    while True:
        # write header
        df_rows, df_cols = crawl_data.shape
        x, y = 1, 1
        df_list = list(crawl_data)          
        while y <= df_cols:
            data.cell( row = x, column = y ).value = df_list[y-1]        
            y += 1
        x += 1
        # write data
        n = 0
        while x <= df_rows+1:               
            y = 1
            df_list = list(crawl_data.loc[n])
            while y <= df_cols:
                data.cell( row = x, column = y ).value = df_list[y-1]
                y += 1
            x += 1
            n += 1
        break
    template.save( fr'L:\Commercial\Operations\Technical SEO\Automation\Technical Healthchecks\03. Data\Crawl Data\{client} - Tech Healthcheck Template.xlsx' )    
    print( '\n- Data Tab Updated' )

# =============================================================================
# update healthcheck tab
# =============================================================================

    df_data = pd.read_excel( fr'L:\Commercial\Operations\Technical SEO\Automation\Technical Healthchecks\03. Data\Crawl Data\{client} - Tech Healthcheck Template.xlsx', sheet_name = 'Data' )
    df_healthcheck = pd.read_excel( fr'L:\Commercial\Operations\Technical SEO\Automation\Technical Healthchecks\03. Data\Crawl Data\{client} - Tech Healthcheck Template.xlsx', sheet_name = 'Healthcheck' )
    df_data = df_data[[ 'Issue', report_month ]]
    df_data = df_data.groupby( 'Issue' )[report_month].sum()
    df_healthcheck = df_healthcheck.merge( df_data, on='Issue', how='left' )
    df_rows, df_cols = df_healthcheck.shape
    x, y = 1, 1

    while True:
        # write header
        df_rows, df_cols = df_healthcheck.shape
        df_list = list(df_healthcheck)          
        while y <= df_cols:
            healthcheck.cell( row = x, column = y ).value = df_list[y-1]
            y += 1
        x += 1
        # write data
        n = 0
        while x <= df_rows+1:
            y = 1
            df_list = list(df_healthcheck.iloc[n])
            while y <= df_cols:
                healthcheck.cell( row = x, column = y ).value = df_list[y-1]
                y += 1
            x += 1
            n += 1
        break
    print( '- Healthcheck Tab Updated' )
    
# save over template file
        
    template.save( fr'L:\Commercial\Operations\Technical SEO\Automation\Technical Healthchecks\03. Data\Crawl Data\{client} - Tech Healthcheck Template.xlsx' )    
    print( f'\n{client} Template Saved' 
            '\n'
            '\n*   *   *   *   *   *   *   *   *   *' )
#    break

print( '\n-----------------------------' 
       '\nHealthcheck Templates Updated'
       '\n-----------------------------' )






'''
# =============================================================================
# generate healthcheck for each client
# =============================================================================
'''






# create path if does not exist

# =============================================================================
# if not os.path.exists( fr'L:\Commercial\Operations\Technical SEO\Technical Healthchecks\01. Healthchecks\{year}' ):
#     os.mkdir( fr'L:\Commercial\Operations\Technical SEO\Technical Healthchecks\01. Healthchecks\{year}' )
# if not os.path.exists( fr'L:\Commercial\Operations\Technical SEO\Technical Healthchecks\01. Healthchecks\{year}\{month}' ):
#     os.mkdir( fr'L:\Commercial\Operations\Technical SEO\Technical Healthchecks\01. Healthchecks\{year}\{month}')   
# =============================================================================

# iterate through rows of healthcheck list and produce one healthcheck per client

for i in hc_list.index:
    client = hc_list['Client'][i]
    project_id = hc_list['Project ID'][i]
    brand = hc_list[ 'Brand' ][i]
        
# load workbook as openpyxl and convert to dataframe 
# (gets around an issue where pandas wouldn't load the workbook properly)
    
    print( f'\nOpening {client} Healthcheck')
    try:
        df = openpyxl.load_workbook( filename=fr'L:\Commercial\Operations\Technical SEO\Automation\Technical Healthchecks\03. Data\Crawl Data\{client} - Tech Healthcheck Template.xlsx', data_only=True )
        df = df[ 'Healthcheck' ]
        df = pd.DataFrame( df.values )
        df_rows, df_cols = df.shape
        xl_rows = df_rows + 4
        print( f'Starting to Write {client} Healthcheck' )
    except FileNotFoundError:
        print( f'{client} Healthcheck Template not found' )
        continue    
    df_rows, df_cols = df.shape
    xl_rows = df_rows+4
    df = df.fillna( '' )

# create blank workbook with xlsxwriter

#    wb = xl.Workbook( fr'L:\Commercial\Operations\Technical SEO\Technical Healthchecks\01. Healthchecks\{year}\{month}\{client} - Tech Healthcheck.xlsx' )
    wb = xl.Workbook( fr'L:\Commercial\Operations\Technical SEO\Technical Healthchecks\01. Healthchecks\{client} - Tech Healthcheck.xlsx' )
    ws = wb.add_worksheet( 'Healthcheck' )
    
# column widths

    ws.set_column( 0, 0, 5 )
    ws.set_column( 1, 1, 45 )
    ws.set_column( 2, 2, 90 )
    ws.set_column( 3, 3, 60 )

# brand colours

    ipro_colour = '#8dc63f'
    ipro_spark = '#066bb1'

    carat_colour = '#00beff'
    carat_spark = '#4ee6C6'

    vizeum_colour = '#fac600'
    vizeum_spark = '#f6861f'

    text = '#636363'

# additional formatting

    ws.freeze_panes(6, 2)                                                      # Freeze first 6 rows and first 2 columns.
    ws.set_zoom(70)                                                            # Set zoom to 70%
    ws.hide_gridlines(2)                                                       # Hide Gridlines
#    ws.hide_row_col_headers()                                                  # hide headers
    
#  CELL FORMATTING

# heading formatting (conditional on brand client is with)

    if brand == 'Carat':
        client_name = wb.add_format({ 'bold':True, 'indent':0, 'font_color':text, 'font_size':20, 'align':'center', 'valign':'vcenter', 'rotation':0 })
        l_head = wb.add_format({ 'bold':False, 'indent':0, 'bg_color':carat_colour, 'font_color':carat_colour, 'font_size':12, 'align':'center', 'valign':'vcenter', 'rotation':0, 'border_color':text, 'top':1, 'bottom':1, 'left':1, 'right':0 })
        c_head = wb.add_format({ 'bold':False, 'indent':0, 'bg_color':carat_colour, 'font_color':carat_colour, 'font_size':12, 'align':'center', 'valign':'vcenter', 'rotation':0, 'border_color':text, 'top':1, 'bottom':1, 'left':0, 'right':0 })
        r_head = wb.add_format({ 'bold':False, 'indent':0, 'bg_color':carat_colour, 'font_color':'#ffffff', 'font_size':12, 'align':'center', 'valign':'vcenter', 'rotation':90, 'border_color':text, 'top':1, 'bottom':1, 'left':1, 'right':1, 'num_format':'mmm-yy' })

    if brand == 'Vizeum':
        client_name = wb.add_format({ 'bold':True, 'indent':0, 'font_color':text, 'font_size':20, 'align':'center', 'valign':'vcenter', 'rotation':0 })
        l_head = wb.add_format({ 'bold':False, 'indent':0, 'bg_color':vizeum_colour, 'font_color':vizeum_colour, 'font_size':12, 'align':'center', 'valign':'vcenter', 'rotation':0, 'border_color':text, 'top':1, 'bottom':1, 'left':1, 'right':0 })
        c_head = wb.add_format({ 'bold':False, 'indent':0, 'bg_color':vizeum_colour, 'font_color':vizeum_colour, 'font_size':12, 'align':'center', 'valign':'vcenter', 'rotation':0, 'border_color':text, 'top':1, 'bottom':1, 'left':0, 'right':0 })
        r_head = wb.add_format({ 'bold':False, 'indent':0, 'bg_color':vizeum_colour, 'font_color':'#ffffff', 'font_size':12, 'align':'center', 'valign':'vcenter', 'rotation':90, 'border_color':text, 'top':1, 'bottom':1, 'left':1, 'right':1, 'num_format':'mmm-yy' })

    if brand not in [ 'Carat', 'Vizeum' ]:
        client_name = wb.add_format({ 'bold':True, 'indent':0, 'font_color':text, 'font_size':20, 'align':'center', 'valign':'vcenter', 'rotation':0 })
        l_head = wb.add_format({ 'bold':False, 'indent':0, 'bg_color':ipro_colour, 'font_color':ipro_colour, 'font_size':12, 'align':'center', 'valign':'vcenter', 'rotation':0, 'border_color':text, 'top':1, 'bottom':1, 'left':1, 'right':0 })
        c_head = wb.add_format({ 'bold':False, 'indent':0, 'bg_color':ipro_colour, 'font_color':ipro_colour, 'font_size':12, 'align':'center', 'valign':'vcenter', 'rotation':0, 'border_color':text, 'top':1, 'bottom':1, 'left':0, 'right':0 })
        r_head = wb.add_format({ 'bold':False, 'indent':0, 'bg_color':ipro_colour, 'font_color':'#ffffff', 'font_size':12, 'align':'center', 'valign':'vcenter', 'rotation':90, 'border_color':text, 'top':1, 'bottom':1, 'left':1, 'right':1, 'num_format':'mmm-yy' })

# table Formatting

    sub1 = wb.add_format({ 'bold':True, 'indent':1, 'font_color':text, 'font_size':18, 'align':'left', 'valign':'vcenter', 'rotation':0, 'border_color':text, 'top':1, 'bottom':1, 'left':0, 'right':0 })
    sub2 = wb.add_format({ 'bold':True, 'indent':1, 'font_color':'#ffffff', 'font_size':18, 'align':'left', 'valign':'vcenter', 'rotation':0, 'border_color':text, 'top':1, 'bottom':1, 'left':0, 'right':0 })
    issue = wb.add_format({ 'bold':True, 'indent':1, 'font_color':text, 'font_size':12, 'align':'left', 'valign':'vcenter', 'rotation':0, 'border_color':text, 'top':1, 'bottom':1, 'left':1, 'right':1 })
    detail = wb.add_format({ 'bold':False, 'indent':1, 'font_color':text, 'font_size':12, 'align':'left', 'valign':'vcenter', 'rotation':0, 'border_color':text, 'top':1, 'bottom':1, 'left':1, 'right':1 })
    spark = wb.add_format({ 'bold':False, 'indent':1, 'font_color':'#ffffff', 'font_size':12, 'align':'left', 'valign':'vcenter', 'rotation':0, 'border_color':text, 'top':1, 'bottom':1, 'left':1, 'right':1 })
    data = wb.add_format({ 'bold':False, 'indent':0, 'font_color':text, 'font_size':12, 'align':'center', 'valign':'vcenter', 'rotation':0, 'border_color':text, 'top':1, 'bottom':1, 'left':1, 'right':1 })

# row groups for adding formatting

    xl_head = [ 5 ]
    xl_sub = [ 6, 14, 21, 28, 37, 42, 46 ]
    xl_data = [ 7, 8, 9, 10, 11, 12, 13, 15, 16, 17, 18, 19, 20, 22, 23, 24, 25, 26, 27, 29, 30, 31, 32, 33, 34, 35, 36, 38, 39, 40, 41, 43, 44, 45, 47, 48, 49, 50, 51, 52, 53 ]
    
# format + fill table header
    
    ws.set_row( 2, 30 )
    ws.write(2, 2, f'{client} Healthcheck', client_name)
#    print('1')
# format + fill table data
    
    x = 5
    n = 0 #for df.loc[n]
    while x <= xl_rows:
#        print('2')
        if x in xl_head:
#            print('3')
            y = 1
            df_list = list(df.iloc[n])
            ws.set_row ( x, 60 )
            while y <= df_cols:
                if y == 1:
                    ws.write( x, y, df_list[y-1], l_head)
                    y += 1
                if y == 2:
                    ws.write( x, y, df_list[y-1], c_head)
                    y += 1
                if y == 3:
                    ws.write( x, y, '', c_head)
                    y += 1
                if y in range ( 4, df_cols ):
                    ws.write( x, y, df_list[y-1], r_head)
                    y += 1
                if y == df_cols:
                    ws.write( x, y, df_list[y-1], r_head)
                    y += 1
                    x += 1
                    n += 1
        if x in xl_sub: 
#            print('4')
            y = 1
            df_list = list(df.iloc[n])
            ws.set_row( x, 60 )
            while y <= df_cols:
                if y == 1:
                    ws.write( x, y, df_list[y-1], sub1)
                    y += 1
                if y in range(2, df_cols):
                    ws.write( x, y, df_list[y-1], sub2)
                    y += 1
                if y == df_cols:
                    ws.write( x, y, df_list[y-1], sub2)
                    y += 1
                    x += 1
                    n += 1
        if x in xl_data:
#            print('5')
            y = 1
            df_list = list(df.iloc[n])
            ws.set_row( x, 30 )
            while y <= df_cols:
                if y == 1:
                    ws.write( x, y, df_list[y-1], issue)
                    y += 1
                if y == 2:
                    ws.write( x, y, df_list[y-1], detail)
                    y += 1
                if y == 3:
                    ws.write( x, y, '', spark)
                    y += 1
                if y in range ( 4, df_cols ):
                    ws.write( x, y, df_list[y-1], data)
                    y += 1
                if y == df_cols:
                    ws.write( x, y, df_list[y-1], data)
                    y += 1
                    x += 1
                    n += 1        

# Add sparklines

    x, y = 7, 3

    if brand == 'Carat':
        while x <= xl_rows:
            sparkline = xl_range( x, 4, x, df_cols)
            ws.add_sparkline( x, y, { 'range': sparkline, 'series_color': carat_spark, 'markers':True } )
            x += 1
        ws.insert_image('B2', fr'L:\Commercial\Operations\Technical SEO\Automation\Technical Healthchecks\03. Data\Branding\Logos\Carat Logo.png', { 'y_offset' : 5, 'x_scale' : 0.13, 'y_scale': 0.13 } )
        ws.set_tab_color(carat_colour)

    if brand == 'Vizeum':
        while x <= xl_rows:
            sparkline = xl_range( x, 4, x, df_cols)
            ws.add_sparkline( x, y, { 'range': sparkline, 'series_color': vizeum_spark, 'markers':True } )
            x += 1
        ws.insert_image('B1', fr'L:\Commercial\Operations\Technical SEO\Automation\Technical Healthchecks\03. Data\Branding\Logos\Vizeum Logo.png', { 'x_offset' : 20, 'y_offset' : 10, 'x_scale' : 0.4, 'y_scale': 0.4 } )
        ws.set_tab_color(vizeum_colour)
    if brand not in [ 'Carat', 'Vizeum' ]:
        while x <= xl_rows:
            sparkline = xl_range( x, 4, x, df_cols)
            ws.add_sparkline( x, y, { 'range': sparkline, 'series_color': ipro_spark, 'markers':True } )
            x += 1
        ws.insert_image('B2', fr'L:\Commercial\Operations\Technical SEO\Automation\Technical Healthchecks\03. Data\Branding\Logos\iPro Logo.png', {'x_scale' : 0.22, 'y_scale': 0.22 } )
        ws.set_tab_color(ipro_colour)   

# hide columns

    if df_cols > 9:
        ws.set_column( 4, df_cols-6, 0 )
     
    wb.close()
    print( f'{client} Healthcheck Done'
           '\n'
           '\n*   *   *   *   *   *   *   *   *   *' )
#    break

print( 'DURN!')