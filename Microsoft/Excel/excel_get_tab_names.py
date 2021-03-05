#%% excel_get_tab_names.py

from openpyxl import load_workbook

path = r'C:\Users\JLee35\OneDrive - Dentsu Aegis Network\Python\Trainline\TL SEOmonitor kws historic rankings exports.xlsx'

wb = load_workbook(filename = path, read_only=True)
wb = wb.sheetnames

print('Loaded worksheet names!')

# %% remove completed tabs from the list (wb)

import os

completed = []

for filename in os.listdir(fr'C:\Users\JLee35\OneDrive - Dentsu Aegis Network\Python\Trainline'):
    if filename.endswith(".csv"):
        if filename.startswith('~$'): # skips hidden temp files
            continue
        else:
            account = filename.replace('.csv','')
            completed.append(account)

for c in completed:
    wb.remove(c)

print('Removed completed worksheets!')