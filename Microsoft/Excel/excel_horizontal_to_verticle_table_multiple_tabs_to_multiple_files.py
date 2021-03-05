# excel_horizontal_to_vertical_table_multiple_tabs.py

#%% load tab names

from openpyxl import load_workbook

path = r'C:\Users\JLee35\OneDrive - Dentsu Aegis Network\Python\Trainline\TL SEOmonitor kws historic rankings exports.xlsx'

wb = load_workbook(filename = path, read_only=True)
wb = wb.sheetnames

print('Loaded worksheet names!')

# %% remove completed tabs from the list (wb)

import os

completed = []

for filename in os.listdir(fr'C:\Users\JLee35\OneDrive - Dentsu Aegis Network\Python\Trainline'):
    if filename.endswith(".csv"):
        if filename.startswith('~$'): # skips hidden temp files
            continue
        else:
            account = filename.replace('.csv','')
            completed.append(account)

for c in completed:
    wb.remove(c)

print('Removed completed worksheets!')

#%%

import pandas as pd

for i in wb:
    print(f'loading {i}')
    df = pd.read_excel(path, sheet_name=i)

    n = 6
    df2 = df.iloc[:, :n] # this selects the first n columns
    dates = df.iloc[:, n:].columns.to_list() # this selects all columns after (not including) n

    df_all = pd.DataFrame()
    n=0
    for date in dates:
        ranks = df[date]
        temp = df2
        temp['Rank'] = ranks
        temp['Date'] = date
        temp['Site'] = i
        df_all = df_all.append(temp,ignore_index=True)
        print(f'Completed {n+1} of {len(dates)}')
        n += 1

    df_all = df_all.loc[:,~df_all.columns.str.contains(['2019', '2020', '2021'], case=False)] # remove columns containing '2019' in label
    print('\nSaving...')
    df_all.to_csv(fr'C:\Users\JLee35\OneDrive - Dentsu Aegis Network\Python\Trainline\{i}.csv', index=False)

print('\nDURN!')

# %%
