#%% excel_remove_col_containing_dates.py

import os
import pandas as pd
from openpyxl import load_workbook

path = r'C:\Users\JLee35\OneDrive - Dentsu Aegis Network\Python\Trainline\TL SEOmonitor kws historic rankings exports.xlsx'

wb = load_workbook(filename = path, read_only=True)
wb = wb.sheetnames

print('Loaded worksheet names!')

#%%
path = r'C:\Users\JLee35\OneDrive - Dentsu Aegis Network\Python\Trainline'

for i in wb:
    print(f'loading {i}')
    df = pd.read_csv(fr'{path}\{i}.csv')
        for string in ['2018', '2019', '2020', '2021']:
            df_all = df_all[df_all.columns.drop(list(df_all.filter(regex=string)))] # drop columns containing 'item'
    df.to_csv(fr'{path}\{i}.csv')

# %%
