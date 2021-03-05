#%% excel_detox_accent_characters.py

# load list of tabs in excel file

from openpyxl import load_workbook

path = r'C:\Users\JLee35\OneDrive - Dentsu Aegis Network\Python\Trainline\Decontaminated\Book1.xlsx'

wb = load_workbook(filename = path, read_only=True)
ws = wb.sheetnames
wb.close()

print('Loaded worksheet names!')

#%% remove completed from wb:

import os

completed = []

for filename in os.listdir(fr'C:\Users\JLee35\OneDrive - Dentsu Aegis Network\Python\Trainline\Decontaminated'):
    if filename.endswith(".xlsx"):
        if filename.startswith('~$'): # skips hidden temp files
            continue
        else:
            account = filename.replace('.xlsx','')
            completed.append(account)

completed.remove('Book1')

for c in completed:
    ws.remove(c)

print('Removed completed worksheets!')

# %%

import pandas as pd

for i in ws:
    print(f'{i} loading...')
    df = pd.read_excel(path, sheet_name=i)
    print(f'{i} loaded!')
    df['Keyword'] = df['Keyword'].replace('ÃŸ', 'ß', regex=True)
    df['Keyword'] = df['Keyword'].replace('Ã¤', 'ä', regex=True)
    df['Keyword'] = df['Keyword'].replace('Ã¬', 'ì', regex=True)
    df['Keyword'] = df['Keyword'].replace('Ãª', 'ê', regex=True)
    df['Keyword'] = df['Keyword'].replace('Ã¹', 'ù', regex=True)
    df['Keyword'] = df['Keyword'].replace('Ã ', 'à', regex=True)
    df['Keyword'] = df['Keyword'].replace('Ã®', 'î', regex=True)
    df['Keyword'] = df['Keyword'].replace('Ã»', 'û', regex=True)
    df['Keyword'] = df['Keyword'].replace('Ã¨', 'è', regex=True)
    df['Keyword'] = df['Keyword'].replace('Ã¶', 'ö', regex=True)
    df['Keyword'] = df['Keyword'].replace('Ã§', 'ç', regex=True)
    df['Keyword'] = df['Keyword'].replace('Ã©', 'é', regex=True)
    df['Keyword'] = df['Keyword'].replace('Ãº', 'ú', regex=True)
    df['Keyword'] = df['Keyword'].replace('Ã¡', 'á', regex=True)
    df['Keyword'] = df['Keyword'].replace('Ã³', 'ó', regex=True)
    df['Keyword'] = df['Keyword'].replace('Ã±', 'ñ', regex=True)
    df['Keyword'] = df['Keyword'].replace('Ã¼', 'ü', regex=True)
    df['Keyword'] = df['Keyword'].replace('ƒô', 'ę', regex=True)
    df['Keyword'] = df['Keyword'].replace('≈õ', 'ś', regex=True)
    df['Keyword'] = df['Keyword'].replace('ƒá', 'ć', regex=True)
    df['Keyword'] = df['Keyword'].replace('√≥', 'ó', regex=True)
    df['Keyword'] = df['Keyword'].replace('≈Ç', 'ł', regex=True)
    df['Keyword'] = df['Keyword'].replace('ƒÖ', 'ą', regex=True)
    df['Keyword'] = df['Keyword'].replace('≈º', 'ż', regex=True)
    df['Keyword'] = df['Keyword'].replace('≈Ñ', 'ń', regex=True)
    df['Keyword'] = df['Keyword'].replace('√º', 'ü', regex=True)
    df['Keyword'] = df['Keyword'].replace('Ã­', 'í', regex=True) # though this looks like a standard 'Ã­', it's actually different
    print(f'{i} saving...')
    df.to_excel(fr'C:\Users\JLee35\OneDrive - Dentsu Aegis Network\Python\Trainline\Decontaminated\{i}.xlsx', encoding='utf-8-sig', index=False)
    print(f'{i} saved!')
print('DURN!')

# %%
