#%% excel_remove_col_containing_dates.py

import os
import pandas as pd
from openpyxl import load_workbook

path = r'C:\Users\JLee35\OneDrive - Dentsu Aegis Network\Python\Trainline\TL SEOmonitor kws historic rankings exports.xlsx'

wb = load_workbook(filename = path, read_only=True)
wb = wb.sheetnames

print('Loaded worksheet names!')

#%%
path = r'C:\Users\JLee35\OneDrive - Dentsu Aegis Network\Python\Trainline'

for i in wb:
    print(f'loading {i}')
    df = pd.read_csv(fr'{path}\{i}.csv')
    df = df.loc[:,~df.columns.str.contains('2019', case=False)] # remove columns containing 'unnamed' in label
    df.to_csv(fr'{path}\{i}.csv')

# %%
path = r'C:\Users\JLee35\OneDrive - Dentsu Aegis Network\Python\Trainline'

for i in wb:
    print(f'loading {i}')
    df = pd.read_excel(r'C:\Users\JLee35\OneDrive - Dentsu Aegis Network\Python\Trainline\TL SEOmonitor kws historic rankings exports.xlsx', sheet_name=i)
    df['Keyword'] = df['Keyword'].replace(['Ã¶', 'ÃƒÂ¶'], 'ö')
    df.to_csv(fr'{path}\Decontaminated\{i}.csv', index=False)

#%%

    df = df.loc[:,~df.columns.str.contains(['2019','2020','2021'], case=False)] # remove columns containing '2019' in label
'ÃŸ', 'ß'
'Ã¤', 'ä'
'Ã¬', 'ì'
'Ãª', 'ê'
'Ã¹', 'ù'
'Ã ', 'à'
'Ã®', 'î'
'Ã»', 'û'
'Ã¨', 'è'
'Ã¶', 'ö'
'Ã§', 'ç'
'Ã©', 'é'
'Ãº', 'ú'
'Ã¡', 'á'
'Ã³', 'ó'
'Ã±', 'ñ'
'Ã¼', 'ü'
'ƒô', 'ę'
'≈õ', 'ś'
'ƒá', 'ć'
'√≥', 'ó'
'≈Ç', 'ł'
'ƒÖ', 'ą'
'≈º', 'ż'
'≈Ñ', 'ń'
'√º', 'ü'
'Ã­', 'í'
