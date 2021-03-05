# excel_horizontal_to_vertical_table_multiple_tabs_to_multiple_files.py

#%% load tab names

from openpyxl import load_workbook

file_with_multiple_tabs = r'C:\Users\JLee35\OneDrive - Dentsu Aegis Network\Python\Trainline\TL SEOmonitor kws historic rankings exports.xlsx'
path_to_save = r'C:\Users\JLee35\OneDrive - Dentsu Aegis Network\Python\Trainline'

wb = load_workbook(filename = path, read_only=True)
wb = wb.sheetnames

print('Loaded worksheet names!')

# %% remove completed tabs from the list (wb)

import os

completed = []

for filename in os.listdir(fr'{path_to_save}'):
    if filename.endswith(".csv"):
        if filename.startswith('~$'): # skips hidden temp files
            continue
        else:
            name = filename.replace('.csv','')
            completed.append(name)

for c in completed:
    wb.remove(c)

print('Removed completed worksheets!')

#%%

import pandas as pd

for i in wb:
    print(f'loading {i}')
    df = pd.read_excel(path, sheet_name=i)
    print('\nSaving...')
    df.to_csv(fr'{path_to_save}\{i}.csv', index=False)

print('\nDURN!')

# %%
